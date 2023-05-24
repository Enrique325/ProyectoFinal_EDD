import pandas as pd
import tkinter as tk
from openpyxl import load_workbook
import arbolPromedio as arbolPromedio
import arbolNombres as arbolNombres
import arbolProfesion as arbolProfesion


workbook = load_workbook(filename='C:/Users/52999/OneDrive/Desktop/ESTRUCTURA DE DATOS V2/PROYECTO/EDD/Egresados.xlsx')
df = pd.read_excel("C:/Users/52999/OneDrive/Desktop/ESTRUCTURA DE DATOS V2/PROYECTO/EDD/Egresados.xlsx")


hoja = workbook['Egresados']

def centrar_ventana(ventana):
    ventana.update_idletasks()
    ancho = ventana.winfo_width()
    altura = ventana.winfo_height()
    x = (ventana.winfo_screenwidth() // 2) - (ancho // 2)
    y = (ventana.winfo_screenheight() // 2) - (altura // 2)
    ventana.geometry(f"{ancho}x{altura}+{x}+{y}")


def mostrar_datos():
    
    datos=[]
    
    if var.get() == 0:
        indices = arbolPromedios.listar_ascendente()
    else:
        indices = arbolPromedios.listar_descendente()
    
    #print(indices_ascendente)
    if indices is not None:
        for indice in indices:
            if isinstance(indice, int) and indice <= len(df):
                fila = df.iloc[indice-1]
                datos.append({
                    "Nombre del Alumno": fila["Nombre del Alumno"],
                    "Profesión": fila["Profesión"],
                    "Calificación": fila["Promedio"]
                })
    
    for dato in datos:
        Calificación="{:.2f}".format(dato['Calificación'])
        resultado_text.insert(tk.END, f"Nombre del Alumno: {dato['Nombre del Alumno']}\n")
        resultado_text.insert(tk.END, f"Calificación: {Calificación}\n")
        resultado_text.insert(tk.END, f"Profesión: {dato['Profesión']}\n")
        resultado_text.insert(tk.END, f"---------------------------------\n")

def buscarDato(datos_buscar):
        if datos_buscar is not None:
            datos_interseccion = []
            for indice in datos_buscar:
                fila = df.iloc[indice-1]
                datos = {
                    "Calificación": fila["Promedio"],
                    "Profesión": fila["Profesión"],
                    "Nombre del Alumno": fila["Nombre del Alumno"]
                }
                datos_interseccion.append(datos)
        resultado_text.delete(1.0, tk.END)
        mostrar_datos_label(datos_interseccion)
        
        
        
def mostrar_datos_label(datos_interseccion):
        
    if datos_interseccion:
        if var.get() == 0:
            datos_interseccion.sort(key=lambda x: x["Calificación"])
        else:
            datos_interseccion.sort(key=lambda x: x["Calificación"], reverse=True)

        resultado_text.delete(1.0, tk.END)

        for datos in datos_interseccion:
            Calificación = "{:.2f}".format(datos['Calificación'])
            resultado_text.insert(tk.END, f"Nombre del Alumno: {datos['Nombre del Alumno']}\n")
            resultado_text.insert(tk.END, f"Calificación: {Calificación}\n")
            resultado_text.insert(tk.END, f"Profesión: {datos['Profesión']}\n")
            resultado_text.insert(tk.END, "-----------------\n")
        
        


def obtener_datos():
   
    dato_string = profesion_entry.get()
    dato_entero = calificacion_entry.get()

    try:
        dato_entero = int(dato_entero)
        
        print("Número ingresado:", dato_entero)
    except ValueError:
        print("Ingresa un número válido")
        
    if dato_string and dato_entero:
        egresados_profesion = arbolProfesion.buscar_por_profesion(dato_string)
        egresados_calificacion = arbolPromedios.buscar_por_calificacion(dato_entero)
        indices_interseccion = egresados_calificacion.intersection(egresados_profesion)
        
        print(indices_interseccion)
        
        buscarDato(indices_interseccion)
                
    elif dato_string:
            egresados_profesion = arbolProfesion.buscar_por_profesion(dato_string)
            print(egresados_profesion)
            buscarDato(egresados_profesion)
                
    elif dato_entero:
            egresados_calificacion=arbolPromedios.buscar_por_calificacion(dato_entero)
            print(egresados_calificacion)
            buscarDato(egresados_calificacion)

    else:
        #resultado_text.delete(1.0, tk.END)
        resultado_text.insert(tk.END, f"No hay datos")


root = tk.Tk()
root.title("Búsqueda de Egresados")
root.geometry("500x550")
   
profesion_label = tk.Label(root, text="Profesión a buscar:")
profesion_label.pack()

profesion_entry = tk.Entry(root)
profesion_entry.pack()

calificacion_label = tk.Label(root, text="Calificación a buscar:")
calificacion_label.pack()

calificacion_entry = tk.Entry(root)
calificacion_entry.pack()

Ordenamiento = tk.Label(root, text="Selecciona el ordenamiento:")
Ordenamiento.pack()

var = tk.IntVar()
rb_ascendente = tk.Radiobutton(root, text="Ascendente", variable=var, value=0)
rb_ascendente.pack()

rb_descendente = tk.Radiobutton(root, text="Descendente", variable=var, value=1)
rb_descendente.pack()

buscar_button = tk.Button(root, text="Buscar", command=obtener_datos)
buscar_button.pack()

resultado_text = tk.Text(root, height=20)
resultado_text.pack()




# Crear los árboles y cargar los datos

if __name__ == '__main__':
    arbolPromedios = arbolPromedio.ArbolAVL()
    arbolesNombres = arbolNombres.ArbolAVL()
    arbolProfesion = arbolProfesion.ArbolAVL()
    centrar_ventana(root)
    
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        indice = fila[0]
        dato = fila[3]
        if indice is not None and dato is not None:
            arbolPromedios.insertar(int(dato), int(indice))

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        indice = fila[0]
        dato = fila[2]
        if indice is not None and dato is not None:
            arbolProfesion.insertar(dato, int(indice))

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        indice = fila[0]
        dato = fila[1]
        if indice is not None and dato is not None:
            arbolesNombres.insertar(dato, int(indice))
    
    mostrar_datos()

    root.mainloop()

workbook.close()